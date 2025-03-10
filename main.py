import os
os.environ["PROTOCOL_BUFFERS_PYTHON_IMPLEMENTATION"] = "python"  #  To run Paddle OCR in python based protobuf implementation as it has dependency conflict with protobuf package

import cv2
import numpy as np
import pandas as pd
from ultralytics import YOLO
from PyQt6.QtWidgets import QApplication, QProgressBar, QLabel, QMenuBar, QMenu,QPushButton, QVBoxLayout, QWidget, QFileDialog, QMessageBox, QGroupBox, QHBoxLayout, QFrame
from PyQt6.QtGui import QPixmap,QAction, QImage
from PyQt6.QtCore import Qt, QThread, pyqtSignal
import mediapipe as mp
import time
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from paddleocr import PaddleOCR



mp_face_detection = mp.solutions.face_detection
mp_drawing = mp.solutions.drawing_utils

IOU_THRESHOLD = 0.2  # Adjust based on your use case
BUFFER = 5  # Allows a small buffer for intersection calculation
FALLBACK_HEADBOX = 0.2
WIOU_THRESHOLD=0.3
MP_FACE_MODEL=0
MP_MIN_DETECTION_CONFIDENCE=0.5
CAMERA_SOURCE = 1 #Webcam / Camera Source REMEMBER WINDOWS HAS WEBCAM ACCESS ISSUES TO PYTHON

# Load YOLO models
motorcycle_model = YOLO("models/yolov8n.pt")  # Pretrained COCO model (motorcycles)
helmet_model = YOLO("models/helmetYoloV8_25epochs.pt")  # Custom helmet detection model
plate_model = YOLO("models/license_plate_detector.pt")  # Custom license plate detection model

# Initialize OCR
ocr = PaddleOCR(use_angle_cls=True, lang='en')  # Angle correction enabled

# Define Colors
COLOR_MOTORCYCLE = (0, 255, 0)  # Green
COLOR_NO_HELMET = (0, 0, 255)  # Red
COLOR_LICENSE_PLATE = (0, 0, 255)  # Red

#Report Filepath
EXCEL_FILE = "report.xlsx"



def get_head_region(image, px1, py1, px2, py2):
    """ Detects head region using Mediapipe Face Detection instead of assuming 20% height. """
    with mp_face_detection.FaceDetection(MP_FACE_MODEL, MP_MIN_DETECTION_CONFIDENCE) as face_detection:
        image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        results = face_detection.process(image_rgb)

        if results.detections:
            for detection in results.detections:
                bboxC = detection.location_data.relative_bounding_box
                h, w, _ = image.shape
                x1, y1, w_box, h_box = (
                    int(bboxC.xmin * w), int(bboxC.ymin * h),
                    int(bboxC.width * w), int(bboxC.height * h)
                )

                # Ensure the detected face is within the person bounding box
                if x1 >= px1 and x1 + w_box <= px2 and y1 >= py1 and y1 + h_box <= py2:
                    return image[y1:y1 + h_box, x1:x1 + w_box]  # Extract face region

    # Fallback: Use the top 20% bounding box if no face is detected
    head_y1 = py1
    head_y2 = py1 + int((py2 - py1) * FALLBACK_HEADBOX)
    return image[head_y1:head_y2, px1:px2]

def get_center(box):
    """Calculate the center of a bounding box."""
    x1, y1, x2, y2 = box
    return ((x1 + x2) // 2, (y1 + y2) // 2)

def iou(box1, box2):
    """Compute Intersection over Union (IoU) between two bounding boxes."""
    x1, y1, x2, y2 = box1
    x1g, y1g, x2g, y2g = box2

    xi1 = max(x1, x1g) - BUFFER
    yi1 = max(y1, y1g) - BUFFER
    xi2 = min(x2, x2g) + BUFFER
    yi2 = min(y2, y2g) + BUFFER

    inter_area = max(0, xi2 - xi1) * max(0, yi2 - yi1)
    box1_area = (x2 - x1) * (y2 - y1)
    box2_area = (x2g - x1g) * (y2g - y1g)

    # union_area = box1_area + box2_area - inter_area
    union_area = min(box1_area, box2_area)  # Normalize based on smaller object
    return inter_area / union_area if union_area > 0 else 0


def aspect_ratio_similarity(box1, box2):
    """Compute Aspect Ratio Similarity (ARS) between two bounding boxes."""
    x1, y1, x2, y2 = box1
    x1g, y1g, x2g, y2g = box2

    aspect_ratio1 = (x2 - x1) / (y2 - y1 + 1e-6)  # Avoid division by zero
    aspect_ratio2 = (x2g - x1g) / (y2g - y1g + 1e-6)

    return min(aspect_ratio1, aspect_ratio2) / max(aspect_ratio1, aspect_ratio2)  # Ratio similarity


def weighted_iou(box1, box2, alpha=0.7):
    """Compute Weighted IoU (WIoU) by combining IoU and Aspect Ratio Similarity."""
    iou_score = iou(box1, box2)
    ars_score = aspect_ratio_similarity(box1, box2)

    return alpha * iou_score + (1 - alpha) * ars_score  # Weighted combination



def associate_persons_with_motorcycles(persons, motorcycles, threshold , alpha=0.7):
    """Associate each detected person with the most relevant motorcycle using Weighted IoU."""
    associations = {}

    for person in persons:
        best_match = None

        best_score = 0  # WIoU score

        for motorcycle in motorcycles:
            wiou_score = weighted_iou(person, motorcycle, alpha)
            if wiou_score > best_score:
                best_score = wiou_score
                best_match = motorcycle

        # Apply Weighted IoU threshold
        if best_match and best_score >= threshold:
            associations[person] = best_match

    return associations



class VideoProcessor(QThread):
    frame_processed = pyqtSignal(np.ndarray)

    def __init__(self, video_path, parent):
        super().__init__()
        self.video_path = video_path
        self.parent = parent
        self.running = True

    def run(self):
        self.parent.process_video(self.video_path)



class RealTimeProcessor(QThread):
    frame_processed = pyqtSignal(np.ndarray)

    def __init__(self, parent):
        super().__init__()
        self.parent = parent
        self.running = True

    def run(self):
        cap = cv2.VideoCapture(0)  # Open webcam
        time.sleep(1.000)
        if not cap.isOpened():
            print("🔴 Error: Could not open webcam.")
            return

        while self.running:
            ret, frame = cap.read()
            if not ret:
                print("⚠️ Frame not captured.")
                break

            # Process the frame
            processed_frame, _ = self.parent.process_image(frame, is_video=True)

            # Emit signal to update UI
            self.frame_processed.emit(processed_frame)

        cap.release()


def format_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    # Define styles
    bold_font = Font(bold=True, color="FFFFFF")  # White text
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Blue background
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    # Apply formatting to header row
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    # Apply formatting to all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_alignment

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get column letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2  # Add extra padding

    # Save the formatted file
    wb.save(file_path)


class HelmetDetectionApp(QWidget):
    def __init__(self):
        super().__init__()
        # self.image_path = ""
        self.processed_image_path = ""
        self.license_plate_text = ""
        self.media_path = ""

        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Helmet Violation Detector")
        self.setGeometry(200, 200, 950, 700)  # More spacious UI

        # Apply a modern dark theme with soft shadows and smooth UI
        self.setStyleSheet("""
            QWidget {
                background-color: #181818;
                color: #E0E0E0;
                font-family: 'Arial';
            }
            QPushButton {
                background-color: #1E88E5;
                color: white;
                padding: 12px;
                border-radius: 10px;
                font-size: 14px;
                font-weight: bold;
                transition: all 0.3s;
            }
            QPushButton:hover {
                background-color: #1565C0;
                transform: scale(1.05);
            }
            QPushButton:pressed {
                background-color: #0D47A1;
            }
            QProgressBar {
                border: 2px solid #1E88E5;
                border-radius: 5px;
                text-align: center;
                background: #222;
            }
            QProgressBar::chunk {
                background-color: #1E88E5;
                width: 10px;
            }
            QLabel {
                font-size: 14px;
                font-weight: bold;
            }
            QMenuBar {
                background-color: #333;
                color: white;
                padding: 6px;
            }
            QMenuBar::item {
                background-color: transparent;
                padding: 6px 12px;
            }
            QMenuBar::item:selected {
                background-color: #1565C0;
            }
            QMenu {
                background-color: #444;
                color: white;
                border: 1px solid #1565C0;
            }
            QMenu::item:selected {
                background-color: #1565C0;
            }
        """)

        layout = QVBoxLayout()

        # 🌟 Top Menu Bar
        menu_bar = QMenuBar(self)

        # 📂 File Menu
        file_menu = QMenu("📂 File", self)
        upload_action = QAction("Upload Media", self)
        upload_action.triggered.connect(self.upload_media)
        exit_action = QAction("❌ Exit", self)
        exit_action.triggered.connect(self.close)

        file_menu.addAction(upload_action)
        file_menu.addSeparator()
        file_menu.addAction(exit_action)

        # ❓ Help Menu
        help_menu = QMenu("❓ Help", self)
        about_action = QAction("ℹ️ About", self)
        about_action.triggered.connect(self.show_about)

        help_menu.addAction(about_action)

        # Add Menus to the Bar
        menu_bar.addMenu(file_menu)
        menu_bar.addMenu(help_menu)

        layout.setMenuBar(menu_bar)  # Attach to main window

        # 🖼️ Image Display Frame
        self.image_label = QLabel("Upload Media")
        self.image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.image_label.setFrameShape(QFrame.Shape.Box)
        self.image_label.setFrameShadow(QFrame.Shadow.Sunken)
        self.image_label.setStyleSheet("""
            background-color: #222;
            border: 3px solid #1E88E5;
            padding: 15px;
            font-size: 18px;
            font-weight: bold;
            color: #BBDEFB;
            border-radius: 8px;
        """)
        layout.addWidget(self.image_label)

        # 📂 File Upload & Processing Section
        file_group = QGroupBox("📂 Media Controls")
        file_group.setStyleSheet("QGroupBox { font-weight: bold; font-size: 16px; }")
        file_layout = QVBoxLayout()

        upload_btn = QPushButton("📂 Upload Media")
        upload_btn.clicked.connect(self.upload_media)
        file_layout.addWidget(upload_btn)

        analyze_btn = QPushButton("🔍 Analyze Media")
        analyze_btn.clicked.connect(self.analyze_media)
        file_layout.addWidget(analyze_btn)

        report_btn = QPushButton("📊 Create Report")
        report_btn.clicked.connect(self.create_report)
        file_layout.addWidget(report_btn)

        file_group.setLayout(file_layout)
        layout.addWidget(file_group)

        # 🎥 Real-Time Section
        real_time_group = QGroupBox("🎥 Real-Time Detection")
        real_time_group.setStyleSheet("QGroupBox { font-weight: bold; font-size: 16px; }")
        real_time_layout = QHBoxLayout()

        self.real_time_btn = QPushButton("🟢 START Real-Time Analysis")
        self.real_time_btn.setCheckable(True)  # Allow toggling
        self.real_time_btn.clicked.connect(self.start_real_time)
        real_time_layout.addWidget(self.real_time_btn)

        real_time_group.setLayout(real_time_layout)
        layout.addWidget(real_time_group)

        # ⏳ Progress Bar for Video Processing
        self.progress_bar = QProgressBar(self)
        self.progress_bar.setTextVisible(True)
        layout.addWidget(self.progress_bar)

        self.setLayout(layout)


    def show_about(self):
        current_year = datetime.now().year  # ✅ Get current year dynamically

        about_msg = QMessageBox(self)
        about_msg.setWindowTitle("🔍 About Helmet Violation Detector")
        about_msg.setIcon(QMessageBox.Icon.Information)  # ✅ Add an info icon

        # 🎨 Rich HTML Styling for Professional Look
        about_text = f"""
        <h2 style='color:#1E88E5; text-align:center;'>Helmet Violation Detector</h2>
        <p style='text-align:center; font-size:14px; color:#E0E0E0;'>
            <b>Version:</b> 01.1<br>
            <b>Developed by:</b> Bitmutex Technologies<br>
            AI-powered safety monitoring solution
        </p>
        <hr>
        <p style='text-align:center; color:#AAAAAA; font-size:12px;'>
            © {current_year} Bitmutex Technologies. All rights reserved.
        </p>
        """

        about_msg.setText(about_text)  # ✅ Set Rich Text
        about_msg.setStandardButtons(QMessageBox.StandardButton.Ok)  # ✅ Add OK Button
        about_msg.exec()  # ✅ Show the Dialog

    def upload_media(self):
        file_dialog = QFileDialog()
        file_path, _ = file_dialog.getOpenFileName(self, "Select Image/Video", "",
                                                   "Images & Videos (*.png *.jpg *.jpeg *.mp4 *.avi *.mov)")

        if file_path:
            self.media_path = file_path
            self.is_video = file_path.lower().endswith((".mp4", ".avi", ".mov"))

            if self.is_video:
                cap = cv2.VideoCapture(self.media_path)
                ret, frame = cap.read()
                cap.release()
                if ret:
                    self.display_frame(frame)
            else:
                pixmap = QPixmap(self.media_path)
                self.image_label.setPixmap(pixmap.scaled(600, 400))


    def analyze_media(self):
        if not self.media_path:
            QMessageBox.warning(self, "Warning", "Please upload an image or video first.")
            return

        if self.is_video:
            self.video_thread = VideoProcessor(self.media_path, self)
            self.video_thread.start()
        else:
            self.processed_path, self.license_plate_text = self.process_image(self.media_path)
            if self.processed_path:
                pixmap = QPixmap(self.processed_path)
                self.image_label.setPixmap(pixmap.scaled(600, 400))


    def process_video(self, video_path, output_folder="output/"):
        cap = cv2.VideoCapture(video_path)
        output_video_path = os.path.join(output_folder, "processed_" + os.path.basename(video_path))
        os.makedirs(output_folder, exist_ok=True)

        fourcc = cv2.VideoWriter_fourcc(*'mp4v')
        fps = int(cap.get(cv2.CAP_PROP_FPS))
        width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        out = cv2.VideoWriter(output_video_path, fourcc, fps, (width, height))

        frame_count = 0  # Track processed frames
        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))  # Get total frames in video

        while cap.isOpened():
            ret, frame = cap.read()

            if not ret:
                print("🔴 No more frames to read. Exiting video processing loop.")
                break

            frame_count += 1
            print(f"Processing frame {frame_count} / {total_frames}")

            processed_frame, license_plate_text = self.process_image(frame, is_video=True)

            if processed_frame is None:
                print("⚠️ Skipping invalid frame.")
                continue  # Skip if processing failed

            out.write(processed_frame)  # Write processed frame to output video

            # 🔄 Update progress bar
            progress = int((frame_count / total_frames) * 100)
            self.progress_bar.setValue(progress)

            if frame_count >= total_frames:
                print("✅ Processed all frames. Stopping.")
                break  # Ensure loop stops

            #Save to Excel
            self.save_to_excel(video_path, output_video_path, license_plate_text)



        cap.release()
        out.release()
        QMessageBox.information(self, "Processing Complete", f"Processed video saved at {output_video_path}")
        self.display_first_video_frame(output_video_path)



    def display_frame(self, frame):
        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        h, w, ch = frame_rgb.shape
        bytes_per_line = ch * w
        qimage = QImage(frame_rgb.data, w, h, bytes_per_line, QImage.Format.Format_RGB888)
        pixmap = QPixmap.fromImage(qimage)
        self.image_label.setPixmap(pixmap.scaled(600, 400))

    def display_first_video_frame(self, video_path):
        cap = cv2.VideoCapture(video_path)
        ret, first_frame = cap.read()
        cap.release()
        if ret:
            self.display_frame(first_frame)




    def start_real_time(self):
        """Toggle real-time helmet detection from webcam."""
        if hasattr(self, 'real_time_thread') and self.real_time_thread.isRunning():
            # 🛑 Stop the real-time thread
            self.real_time_thread.running = False
            self.real_time_thread.quit()
            self.real_time_thread.wait()
            print("🔴 Stopped real-time detection.")

            # 🔄 Update button to "START"
            self.real_time_btn.setText("🟢 START Real-Time Analysis")
            self.real_time_btn.setStyleSheet("background-color: #1E88E5; color: white;")
            return

        print("✅ Starting real-time detection...")

        # 🟢 Start the real-time detection
        self.real_time_thread = RealTimeProcessor(self)
        self.real_time_thread.frame_processed.connect(self.display_frame)
        self.real_time_thread.start()

        # 🔄 Update button to "STOP"
        self.real_time_btn.setText("🔴 STOP Real-Time Analysis")
        self.real_time_btn.setStyleSheet("background-color: #D32F2F; color: white;")


    def create_report(self):
        if not self.license_plate_text:
            QMessageBox.warning(self, "Warning", "No license plate detected yet!")
            return

        self.save_to_excel(self.media_path, self.processed_path ,self.license_plate_text)

        os.system(f"start {EXCEL_FILE}")  # Open Excel file (Windows)
        QMessageBox.information(self, "Success", "Report saved and opened successfully.")


    def save_to_excel(self, media_path_input, media_path_output, license_plate_text):
        # Format the paths
        output_full_path = os.path.abspath(media_path_output)
        input_full_path = os.path.abspath(media_path_input)

        input_dir = os.path.dirname(input_full_path)  # Input path (without filename)
        input_filename = os.path.basename(input_full_path)  # Input filename

        output_dir = os.path.dirname(output_full_path)  # Output path (now absolute)
        output_filename = os.path.basename(output_full_path)  # Output filename

        # Check if Excel file exists
        if os.path.exists(EXCEL_FILE):
            df = pd.read_excel(EXCEL_FILE, engine="openpyxl")
        else:
            df = pd.DataFrame(
                columns=["Input Path", "Output Path", "Input Filename", "Output Filename", "License Plate Number"])

        # Create new entry
        new_entry = pd.DataFrame([[input_dir, output_dir, input_filename, output_filename, license_plate_text]],
                                 columns=df.columns)

        # Append to DataFrame
        df = pd.concat([df, new_entry], ignore_index=True)

        # Save to Excel
        with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as writer:
            df.to_excel(writer, index=False)

        # Format the Excel file
        format_excel(EXCEL_FILE)

    def process_image(self, image_path, is_video=False, output_folder="output/"):

        # 🖼️ Handle video frames directly
        if is_video:
            image = image_path  # Already a NumPy array (frame)
            image_name = "frame.jpg"  # Placeholder name for video frames
        else:
            image = cv2.imread(image_path)
            if image is None:
                print(f"Error: Could not read image {image_path}")
                return None, ""
            image_name = os.path.basename(image_path)

        license_plate_text = self.license_plate_text

        # 🏍️ Detect motorcycles and persons
        motorcycle_results = motorcycle_model(image, conf=0.4, iou=0.5)
        person_results = motorcycle_model(image, conf=0.3, iou=0.8)

        motorcycles = []
        persons = []

        # 🚀 Extract motorcycle bounding boxes
        for result in motorcycle_results:
            for box in result.boxes:
                x1, y1, x2, y2 = map(int, box.xyxy[0])
                cls = int(box.cls[0].item())
                if cls == 3:  # Motorcycle class in COCO
                    motorcycles.append((x1, y1, x2, y2))

        # 🚶 Extract person bounding boxes
        for result in person_results:
            for box in result.boxes:
                x1, y1, x2, y2 = map(int, box.xyxy[0])
                cls = int(box.cls[0].item())
                if cls == 0:  # Person class in COCO
                    persons.append((x1, y1, x2, y2))

        # 🔗 Associate persons with motorcycles
        associations = associate_persons_with_motorcycles(persons, motorcycles, WIOU_THRESHOLD)

        detected_plates = {}  # Store plates per motorcycle to prevent duplicate processing

        for motorcycle in motorcycles:
            mx1, my1, mx2, my2 = motorcycle
            persons_on_motorcycle = [p for p, m in associations.items() if m == motorcycle]

            if not persons_on_motorcycle:
                print("⚠️ No person found on the motorcycle. Skipping...")
                continue  # 🚫 Skip if no person is found

            # 🏍️ Detect helmet for each person ONCE per motorcycle
            helmet_status = []
            for person in persons_on_motorcycle:
                px1, py1, px2, py2 = person
                head_roi = get_head_region(image, px1, py1, px2, py2)

                # Detect helmet
                helmet_results = helmet_model(head_roi)
                helmet_present = any(helmet_results[0].boxes)
                helmet_status.append(helmet_present)

                # 🖍️ Draw bounding box for person
                color = COLOR_MOTORCYCLE if helmet_present else COLOR_NO_HELMET
                cv2.rectangle(image, (px1, py1), (px2, py2), color, 3)
                text = "Helmet On" if helmet_present else "No Helmet!"
                cv2.putText(image, text, (px1, py1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)

            # ✅ 🚫 **Skip if all persons have helmets**
            if all(helmet_status):
                print("✅ All persons have helmets. Skipping license plate detection...")
                continue

            # 🚗 License Plate Detection (only if at least one person has no helmet)
            if motorcycle in detected_plates:
                print("⚠️ License plate already detected for this motorcycle. Skipping duplicate detection...")
                continue  # 🚫 Skip duplicate processing

            motorcycle_roi = image[my1:my2, mx1:mx2]
            plate_results = plate_model(motorcycle_roi)

            if not plate_results or not any(plate_results[0].boxes):
                print("⚠️ No license plate found for this motorcycle. Skipping...")
                continue  # 🚫 Skip if no plate found

            for plate_result in plate_results:
                for plate_box in plate_result.boxes:
                    px1, py1, px2, py2 = map(int, plate_box.xyxy[0])
                    plate_roi = motorcycle_roi[py1:py2, px1:px2]

                    # Preprocess the license plate region
                    plate_roi_gray = cv2.cvtColor(plate_roi, cv2.COLOR_BGR2GRAY)
                    plate_roi_denoised = cv2.fastNlMeansDenoising(plate_roi_gray, h=10)
                    plate_roi_thresh = cv2.adaptiveThreshold(plate_roi_denoised, 255,
                                                             cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                                             cv2.THRESH_BINARY, 11, 2)
                    plate_roi_resized = cv2.resize(plate_roi_thresh, None, fx=1.5, fy=1.5,
                                                   interpolation=cv2.INTER_CUBIC)

                    # Convert BGR to RGB (PaddleOCR expects RGB)
                    plate_roi_rgb = cv2.cvtColor(plate_roi_denoised, cv2.COLOR_GRAY2RGB)

                    # Perform OCR with PaddleOCR
                    result = ocr.ocr(plate_roi_rgb, cls=True)
                    print("Raw Result:", result)

                    # Extract text from PaddleOCR result
                    # Extract text from PaddleOCR result safely
                    if result and isinstance(result[0], list) and len(result[0]) > 0:
                        license_plate_text = " ".join([entry[1][0] for entry in result[0]])  # Extract text
                        print("🚗 Detected Plate:", license_plate_text)

                        # Store plate once per motorcycle
                        detected_plates[motorcycle] = license_plate_text

                        # Draw bounding box and text on original image
                        cv2.rectangle(image, (mx1 + px1, my1 + py1), (mx1 + px2, my1 + py2), COLOR_LICENSE_PLATE, 2)
                        cv2.putText(image, f"Plate: {license_plate_text}", (mx1 + px1, my1 + py1 - 10),
                                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, COLOR_LICENSE_PLATE, 2)
                    else:
                        print("⚠️ No text detected on the license plate!")


        # 📁 Save processed image for single-image mode
        os.makedirs(output_folder, exist_ok=True)
        output_path = os.path.join(output_folder, f"processed_{image_name}")
        cv2.imwrite(output_path, image)
        # 🎥 Return processed frame for video
        if is_video:
            return image, license_plate_text
        else:
            return output_path, license_plate_text


if __name__ == "__main__":
    app = QApplication([])
    window = HelmetDetectionApp()
    window.show()
    app.exec()